from openpyxl import Workbook, load_workbook
from collections import Counter
import xlrd
from os import listdir, chdir, getcwd, mkdir
from shutil import move, copy

# This function checks if the download_path exist and returns download_path
def readDownloadDir(download_path):
	current_dir = getcwd()

	try:
		chdir(download_path)
		print  "Download_dir set to %s" % (download_path)

	except OSError:
		print "%s could not be opened" % (download_path)

	finally:
		chdir(current_dir)

	return download_path


# Check if the Firefox download dir is clear of .xls files
def isDownloadDirClear(download_dir):
	is_download_dir_clear = True

	dir_entries = listdir(download_dir)
	for entry in dir_entries:
		if entry[-5:] == ".xlsx" or entry [-4:] == ".xls":
			is_download_dir_clear = False

	return is_download_dir_clear

# Check if the dir exists
# If not, create it
def checkMakeDir(dir_path):
	current_dir = getcwd()

	try:
		chdir(dir_path)
		print "%s exists" % (dir_path)
	except OSError:
		mkdir(dir_path)
		print "%s was created" % (dir_path)

	finally:
		chdir(current_dir)

	return

# For report builder
# Create main data structure: a dictionary
# This function reads a .xlsx file which must contain
# 1) Company names, 2) CIQ IDs, 3) Batch no
# Key: Firm_name, Values: list of [company_id, batch_no]
def getCompanyNamesInfo(code_name):
	company_names_info = {}
	cwd = getcwd()
	ids_file_path = cwd + "/firm_lists/" + code_name + ".xlsx"

	# Load workbooks into a dictionary of {'firm_name' : [company id, batch_no]}
	master_table = load_workbook(ids_file_path)
	ws = master_table.active
	print "Workbook %s loaded" % (ids_file_path)

	# Find columns containing the companyname, CIQ ID, and the batch_no
	for col_no in range(1, 20):
		col_title = ws.cell(row=1, column=col_no).value

		if col_title == "companyname":
			company_name_col = col_no
		elif col_title == "excelcompanyid":
			company_id_col = col_no
		elif col_title == "batch_no":
			batch_no_col = col_no

	print "columns: [names, ids, batch_no] = [%d, %d,  %d]"\
	      % (company_name_col,company_id_col, batch_no_col)

	company_names_info = {}

	# Read the company info into memory, stopping when encountering an empty cell in col 1
	# Max number of rows is 1 million
	for row_no in range(2,1000000):
		if ws.cell(row=row_no, column=1).value is None:
			break

		company_name = ws.cell(row = row_no, column = company_name_col).value
		batch_no = int(ws.cell(row = row_no, column = batch_no_col).value)
		company_id = ws.cell(row = row_no, column = company_id_col).value

		company_names_info[company_name] = [company_id, batch_no]

	print "%d company names and info loaded into memory" % (len(company_names_info))

	return company_names_info

# For Report Builder
# From the downloaded file, find out what report type and batch number it is
# Parameters are the downloaded file and the company_names_info dict
def getTrueName(rawfile, company_names_info):
	true_name = "Invalid"

	try:
		excel_file = xlrd.open_workbook(rawfile)
	except IOError:
		exit(rawfile + " not found.")
		sheet_names = excel_file.sheet_names()
		report_type = "unknown"
		batch_no_vote = [] 

		# Iterate across the sheets to guess the batch no
		# Stop when either 4 firms are matched or there are no worksheets left
		for sheet_name in sheet_names:
			# No data => next sheet
			if sheet_name == "No Data":
				batch_no_vote.append(0)
				continue

			try:
				excel_sheet = excel_file.sheet_by_name(sheet_name)
				company_name = excel_sheet.cell(1, 0).value

			# Where the data is out of alignment (unexpected cell arrangement)
			except IndexError:
				batch_no_vote.append(0)
				continue

			# No data, then move to next worksheet
			if company_name == "No Data":
				batch_no_vote.append(0)
				continue

			# Get the company name and report type
			company_name = company_name[:(company_name.index('>') - 1)]
			report_type = excel_sheet.cell(3, 0).value
			if report_type == "  Customers":
				report_type = "customers"
			elif report_type == "  Suppliers":
				report_type = "suppliers"
			elif report_type == "Corporate Tree":
				report_type = "corporateT"

			# Get the batch_no
			try:
				batch_no_vote.append(company_names_info[company_name][1])

			except IndexError:
				print "%s has no matching batch no." % (company_name)
				batch_no_vote.append(0)
			except KeyError:
				print "%s was not found in master list" % (company_name)
				batch_no_vote.append(0)

			# Once 4 firms are positively matched to a batch no, exit the loop 
			if len(batch_no_vote) >= 4:
				break
		
		# Find the most common batch_number
		batch_no = (Counter(batch_no_vote).most_common(1))[0][0]

		# If either report_type or zero data file, return Invalid
		if report_type == "unknown" or batch_no == 0:
			true_name = "Invalid"
		# Else, create true_name accordingly
		else:
			true_name = report_type + "_batch_" + str(batch_no) + ".xls"

	### Situations when the names returns invalid
	# 1. This file contains zero bytes
	except xlrd.biffh.XLRDError:
		true_name = "Invalid"
	# 2. The file could not be opened 
	except IOError:
		true_name = "Invalid"

	return true_name

# For Report Builder
# This file checks if all files for a given relations download are present
# It checks from file #1 to #last_batch
def findMissing(downloaded_files, relations, last_batch):
	missing = []
	
	# Check if all files in the batch sequence exist
	# Record the missing batch numbers
	for no in range(1, last_batch+1):
		file_name = str(relations) + "_batch_" + str(no) + ".xls" 
		try:
			downloaded_files.index(file_name)
		except ValueError:
			missing.append(no)

	return missing

# For Report Builder
# Creates a list of firms by CIQ IQ to be downloaded
def getBatchList(company_names_info, batch_no):
	
	# Batch creation
	batch_list = []

	for company in company_names_info:
		if company_names_info[company][1] == batch_no:
			firm_id = company_names_info[company][0]
			batch_list.append(firm_id)

	print "Batch #%d has been created" % (batch_no)

	return batch_list

# For Report Builder
# In the case of nil return from adding firms to the Report Builder 
# Create a empty dummy file with appropriate names
def createDummyFile(batch_no, report_type):
	if (report_type == "customer"):
		dummy_file_name = "customers_batch_" + str(batch_no) + ".xls"
	elif (report_type == "supplier"):
		dummy_file_name = "suppliers_batch_" + str(batch_no) + ".xls"
	elif (report_type == "corporate_tree"):
		dummy_file_name = "corporateT_batch_" + str(batch_no) + ".xls"
	else:
		dummy_file_name = "unknown_batch_" + str(batch_no) + ".xls"

	copy("C:/Selenium/capitaliq/example_dummy_file.xls",\
	     "C:/Users/faslxkn\Downloads/" + dummy_file_name)

	return dummy_file_name

# For Report Builder
# Generate an expected download file name using
# 1. report_type 2. number of firms AddFirms
def getDownloadName(report_type, valid_firm_count):
	# Produce an expected filename 
	if report_type == "customer":
		download_name = str(valid_firm_count) + "Companies_CompanyCustomers.xls"
	elif report_type == "supplier":
		download_name = str(valid_firm_count) + "Companies_CompanySuppliers.xls"
	elif report_type == "corporate_tree":
		download_name = str(valid_firm_count) + "Companies_CorporateTree.xls"
	else:
	        download_name = str(valid_firm_count) + "Companies.xls"

	return download_name

# For both Report Builder and Company Screening
# Shift all .xls files from one folder to another
# Typically the default Firefox download dir to storage dir
def moveAllExcelFiles(source_dir, destination_dir):
	entries = listdir(source_dir)
	files_moved = 0

	for entry in entries:
		if entry [-5:] == ".xlsx" or entry[-4:] == ".xls":
			try:
				move(source_dir+"/"+ entry, destination_dir +"/"+ entry)
				files_moved += 1
			except IOError:
				print "%s could not be moved to %s" % (entry, destination_dir +"/"+ entry)
				continue

	return files_moved

# For both Report Builder and Company Screening
# Shift all .part files from one folder to another
# Typically the default Firefox download dir to storage dir
def moveAllPartialFiles(source_dir, destination_dir):
	entries = listdir(source_dir)
	files_moved = 0

	for entry in entries:
		if entry [-5:] == ".part":
			try:
				move(source_dir+"/"+ entry, destination_dir +"/"+ entry)
				files_moved += 1
			except IOError:
				print "%s could not be moved" % (entry)
				continue

	return files_moved

# For both Report Builder and Company Screening
# Check if the download is complete by checking the download_dir for .part files
def checkDownloadComplete(download_path):
	entries = listdir(download_path)

	for entry in entries:
		if entry [-5:] == ".part":
			return False

	return True


	



